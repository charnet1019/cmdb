"""Tests for export_service: formula-injection sanitization and field formatting."""
import pytest

from app.services.export_service import (
    _export_csv_stream,
    _export_excel_stream,
    _sanitize_cell,
    format_field_value,
)


@pytest.mark.parametrize("payload", [
    "=cmd|'/c calc'!A1",
    "+1+1",
    "-2+3",
    "@SUM(A1:A2)",
    "\ttab-prefixed",
])
def test_sanitize_cell_neutralizes_formula_trigger_chars(payload):
    result = _sanitize_cell(payload)
    assert result == "'" + payload
    assert result.startswith("'")


def test_sanitize_cell_leaves_normal_strings_untouched():
    assert _sanitize_cell("db-primary") == "db-primary"
    assert _sanitize_cell("") == ""


def test_sanitize_cell_ignores_non_string_values():
    assert _sanitize_cell(42) == 42
    assert _sanitize_cell(None) is None


async def _one_row_gen(row):
    yield row


@pytest.mark.asyncio
async def test_export_csv_stream_sanitizes_malicious_asset_name():
    columns = [("name", "资产名称")]
    buffer, count = await _export_csv_stream(_one_row_gen({"name": "=cmd|'/c calc'!A1"}), columns)
    content = buffer.getvalue().decode("utf-8-sig")
    assert count == 1
    assert "'=cmd|'/c calc'!A1" in content


@pytest.mark.asyncio
async def test_export_excel_stream_sanitizes_malicious_credentials_field():
    columns = [("credentials", "用户名密码")]

    async def gen():
        yield {"credentials": [{"username": "=SUM(A1:A9)", "password": "secret"}]}

    buffer, count = await _export_excel_stream(gen(), columns)
    assert count == 1
    from openpyxl import load_workbook
    wb = load_workbook(buffer)
    ws = wb.active
    cell_value = ws.cell(row=2, column=1).value
    assert cell_value.startswith("'=")


def test_format_field_value_vendor_only_falls_back_to_platform_for_network_category():
    # host/database assets have no vendor field; showing platform under "厂商"
    # for them is a category label mix-up, not a genuine vendor value.
    network_asset = {"category": "network", "vendor": None, "platform": "H3C"}
    assert format_field_value("vendor", network_asset) == "H3C"

    host_asset = {"category": "host", "vendor": None, "platform": "Linux"}
    assert format_field_value("vendor", host_asset) is None
