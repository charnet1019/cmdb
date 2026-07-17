"""Tests for parse_import_file: row-count cap and formula-value resolution
(data_only=True), using minimally-constructed in-memory xlsx workbooks."""
from io import BytesIO

import pytest
from openpyxl import Workbook

from tests.factories import FakeDB

from app.services import import_service
from app.services import template_service


def _build_workbook(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _host_create_row(name):
    # One value per HOST_CREATE_FIELDS column, in order — only the required
    # columns need real values, the rest can be blank.
    values_by_field = {
        "name": name, "platform": "Linux", "internal_address": "10.0.0.1", "credentials": "u:p",
    }
    return [values_by_field.get(field_name, "") for field_name, _label, _required in template_service.HOST_CREATE_FIELDS]


def _host_create_headers():
    return [label for _field_name, label, _required in template_service.HOST_CREATE_FIELDS]


@pytest.mark.asyncio
async def test_parse_import_file_rejects_too_many_rows(monkeypatch):
    monkeypatch.setattr(import_service, "MAX_IMPORT_ROWS", 2)
    rows = [_host_create_row("h1"), _host_create_row("h2"), _host_create_row("h3")]
    content = _build_workbook(_host_create_headers(), rows)

    with pytest.raises(ValueError, match="导入行数不能超过"):
        await import_service.parse_import_file(content, "host", "create", FakeDB())


@pytest.mark.asyncio
async def test_parse_import_file_accepts_row_count_at_the_limit(monkeypatch):
    monkeypatch.setattr(import_service, "MAX_IMPORT_ROWS", 2)
    rows = [_host_create_row("h1"), _host_create_row("h2")]
    content = _build_workbook(_host_create_headers(), rows)

    valid_records, error_records = await import_service.parse_import_file(content, "host", "create", FakeDB())

    assert len(valid_records) + len(error_records) == 2


@pytest.mark.asyncio
async def test_parse_import_file_reads_workbook_with_data_only_true(monkeypatch):
    """openpyxl's load_workbook is called with data_only=True so formula
    cells yield their cached computed value rather than the formula string.
    We can't easily assert on a *formula* cell without a real spreadsheet
    engine to compute its cached value, so this test pins the call itself."""
    captured = {}
    real_load_workbook = import_service.load_workbook

    def spy_load_workbook(*args, **kwargs):
        captured.update(kwargs)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(import_service, "load_workbook", spy_load_workbook)

    rows = [_host_create_row("h1")]
    content = _build_workbook(_host_create_headers(), rows)

    await import_service.parse_import_file(content, "host", "create", FakeDB())

    assert captured.get("data_only") is True
