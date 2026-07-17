"""Tests for template_service.generate_category_template's required-field markers."""
from openpyxl import load_workbook

from app.services import template_service


def test_gpt_create_template_marks_credentials_as_required():
    buffer, _filename = template_service.generate_category_template("gpt", "create")
    wb = load_workbook(buffer)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "*用户名密码" in headers


def test_gpt_create_template_required_fields_all_have_asterisk():
    for field_name, label, required in template_service.GPT_CREATE_FIELDS:
        if required:
            assert label.startswith("*"), f"{field_name} is required but its label lacks '*'"
