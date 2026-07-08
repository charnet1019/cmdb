"""
Asset Import Service
Handles XLSX template generation and file import processing
"""
import re
from io import BytesIO
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete

from app.core.encryption import encrypt_value
from app.models import Asset, Organization, AssetHostRelation, StorageLocation, Credential

# Track org paths auto-created during the current import batch (cleared per-import)
CreatedOrgs: set = set()


def get_created_orgs() -> set:
    """Return and clear the set of org paths auto-created during the last import."""
    orgs = CreatedOrgs.copy()
    CreatedOrgs.clear()
    return orgs

# ─── Field definitions ───────────────────────────────────────────────
# Format: (field_name, display_label, is_required)

HOST_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "外网地址", False),
    ("internal_address", "*内网地址", True),
    ("credentials", "*用户名密码", True),
    ("cpu", "CPU", False),
    ("memory", "内存", False),
    ("system_disk", "系统盘", False),
    ("data_disk", "数据盘", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("oob", "OOB 地址", False),
    ("oob_username", "OOB 用户名", False),
    ("oob_password", "OOB 密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

HOST_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("cpu", "CPU", False),
    ("memory", "内存", False),
    ("system_disk", "系统盘", False),
    ("data_disk", "数据盘", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("oob", "OOB 地址", False),
    ("oob_username", "OOB 用户名", False),
    ("oob_password", "OOB 密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

NETWORK_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("device_type", "*设备类型", True),
    ("vendor", "*厂商", True),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("status", "*状态", True),
    ("notes", "描述", False),
]

NETWORK_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("device_type", "设备类型", False),
    ("vendor", "厂商", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

DATABASE_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("db_type", "*数据库类型", True),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("runs_on", "运行于", False),
    ("storage_locations", "存储位置", False),
    ("version", "版本", False),
    ("namespace", "命名空间", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

DATABASE_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("db_type", "数据库类型", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("runs_on", "运行于", False),
    ("storage_locations", "存储位置", False),
    ("version", "版本", False),
    ("namespace", "命名空间", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CLOUD_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CLOUD_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

WEB_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "*外网地址", True),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

WEB_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

GPT_CREATE_FIELDS = [
    ("name", "*名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "*外网地址", True),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", True),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

GPT_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CATEGORY_FIELDS = {
    "host": {"create": HOST_CREATE_FIELDS, "update": HOST_UPDATE_FIELDS},
    "network": {"create": NETWORK_CREATE_FIELDS, "update": NETWORK_UPDATE_FIELDS},
    "database": {"create": DATABASE_CREATE_FIELDS, "update": DATABASE_UPDATE_FIELDS},
    "cloud": {"create": CLOUD_CREATE_FIELDS, "update": CLOUD_UPDATE_FIELDS},
    "web": {"create": WEB_CREATE_FIELDS, "update": WEB_UPDATE_FIELDS},
    "gpt": {"create": GPT_CREATE_FIELDS, "update": GPT_UPDATE_FIELDS},
}

CATEGORY_NAMES = {
    "host": "主机",
    "network": "网络设备",
    "database": "数据库",
    "cloud": "云服务",
    "web": "网站服务",
    "gpt": "AI 服务",
}

# ─── Category-specific metadata ──────────────────────────────────────

# Fields passed to Asset() constructor (beyond name/category/created_by_id)
CATEGORY_CREATE_FIELDS = {
    "host": [
        "asset_code", "platform", "external_address", "internal_address",
        "model", "serial_number", "cpu", "memory", "system_disk", "data_disk",
        "applicant", "organization_id", "notes", "oob_address", "oob_username",
    ],
    "network": [
        "asset_code", "device_type", "vendor", "model", "serial_number",
        "external_address", "internal_address", "applicant", "organization_id", "notes",
    ],
    "database": [
        "asset_code", "platform", "db_type", "external_address", "internal_address",
        "applicant", "namespace", "organization_id", "notes",
    ],
    "cloud": [
        "asset_code", "platform", "external_address", "internal_address",
        "organization_id", "notes",
    ],
    "web": [
        "asset_code", "platform", "external_address", "internal_address",
        "applicant", "organization_id", "notes",
    ],
    "gpt": [
        "asset_code", "platform", "external_address", "internal_address",
        "applicant", "organization_id", "notes",
    ],
}

# Fields updated via setattr in update mode
CATEGORY_UPDATE_FIELDS = {
    "host": [
        "name", "asset_code", "platform", "external_address", "internal_address",
        "model", "serial_number", "cpu", "memory", "system_disk", "data_disk",
        "applicant", "notes", "oob_address", "oob_username",
    ],
    "network": [
        "name", "asset_code", "device_type", "vendor", "model", "serial_number",
        "external_address", "internal_address", "applicant", "notes",
    ],
    "database": [
        "name", "asset_code", "platform", "db_type", "external_address",
        "internal_address", "applicant", "namespace", "notes",
    ],
    "cloud": [
        "name", "asset_code", "platform", "external_address", "internal_address", "notes",
    ],
    "web": [
        "name", "asset_code", "platform", "external_address",
        "internal_address", "applicant", "notes",
    ],
    "gpt": [
        "name", "asset_code", "platform", "external_address",
        "internal_address", "applicant", "notes",
    ],
}

# Credential type per category
CATEGORY_CRED_TYPE = {
    "host": "password",
    "network": "password",
    "database": "password",
    "cloud": "api_key",
    "web": "password",
    "gpt": "api_key",
}

# ─── Template example data ───────────────────────────────────────────

CATEGORY_EXAMPLES = {
    "host": {
        "title_create": "主机导入模板",
        "title_update": "主机更新模板",
        "create": [
            "test-server-01", "CI001", "Default/研发部/服务器组", "Linux",
            "192.168.1.100", "10.0.0.100", "admin:123456", "8 核", "16GB",
            "500GB SSD", "2TB HDD", "Dell R740", "SN123456", "192.168.1.200",
            "admin", "admin123", "张三", "启用", "测试服务器 - 开发环境",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "test-server-01", "CI001",
            "Default/研发部/服务器组", "Linux", "192.168.1.100", "10.0.0.100",
            "admin:123456", "8 核", "16GB", "500GB SSD", "2TB HDD", "Dell R740",
            "SN123456", "192.168.1.200", "admin", "admin123", "张三", "启用",
            "测试服务器 - 开发环境",
        ],
    },
    "network": {
        "title_create": "网络设备导入模板",
        "title_update": "网络设备更新模板",
        "create": [
            "Core-SW-01", "NW001", "Default/研发部/网络设备", "交换机", "Cisco",
            "C9300-48P", "FCW1234D001", "10.0.0.1", "192.168.1.1",
            "admin:cisco123\nnetadmin:netpass123", "启用", "核心交换机 - 生产环境",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "Core-SW-01", "NW001",
            "Default/研发部/网络设备", "交换机", "Cisco", "C9300-48P", "FCW1234D001",
            "10.0.0.1", "192.168.1.1", "admin:cisco123\nnetadmin:netpass123",
            "启用", "核心交换机 - 生产环境",
        ],
    },
    "database": {
        "title_create": "数据库导入模板",
        "title_update": "数据库更新模板",
        "create": [
            "MySQL-Prod-01", "DB001", "Default/研发部/数据库", "RDS", "MySQL",
            "", "192.168.1.100:3306", "root:mysqlroot123\napp:apppass",
            "web-server-01", "/var/lib/mysql|data|主数据目录", "8.0.32", "main",
            "张三", "True", "生产环境主数据库",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "MySQL-Prod-01", "DB001",
            "Default/研发部/数据库", "RDS", "MySQL", "", "192.168.1.100:3306",
            "root:mysqlroot123\napp:apppass", "web-server-01",
            "/var/lib/mysql|data|主数据目录", "8.0.32", "main", "张三",
            "running", "生产环境主数据库",
        ],
    },
    "cloud": {
        "title_create": "云服务导入模板",
        "title_update": "云服务更新模板",
        "create": [
            "AWS-Prod-Account", "CL001", "Default/研发部/云服务", "AWS",
            "https://aws.amazon.com", "10.0.0.1",
            "AKIAIOSFODNN7EXAMPLE:wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
            "True", "生产环境 AWS 账号",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "AWS-Prod-Account", "CL001",
            "Default/研发部/云服务", "AWS", "https://aws.amazon.com", "10.0.0.1",
            "AKIAIOSFODNN7EXAMPLE:wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
            "启用", "生产环境 AWS 账号",
        ],
    },
    "web": {
        "title_create": "网站服务导入模板",
        "title_update": "网站服务更新模板",
        "create": [
            "Jira", "WB001", "Default/研发部/应用系统", "Nginx",
            "https://jira.example.com", "http://192.168.1.100:8080",
            "admin:jiraadmin\nreadonly:readonly123", "张三", "True",
            "项目管理平台",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "Jira", "WB001",
            "Default/研发部/应用系统", "Nginx", "https://jira.example.com",
            "http://192.168.1.100:8080",
            "admin:jiraadmin\nreadonly:readonly123", "项目管理平台",
        ],
    },
    "gpt": {
        "title_create": "AI 服务导入模板",
        "title_update": "AI 服务更新模板",
        "create": [
            "OpenAI-API", "AI001", "Default/研发部/AI 服务", "OpenAI",
            "https://api.openai.com/v1", "",
            "sk-key:sk-abc123xyz\nclue-key:sk-clue456", "张三", "True",
            "AI 服务 API",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "OpenAI-API", "AI001",
            "Default/研发部/AI 服务", "OpenAI", "https://api.openai.com/v1", "",
            "sk-key:sk-abc123xyz\nclue-key:sk-clue456", "张三", "running",
            "AI 服务 API",
        ],
    },
}

# ─── Shared openpyxl styles ──────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
_HINT_FONT = Font(italic=True, color="666666", size=9)
_THIN_BORDER = Border(
    left=Side(style='thin', color='E7E6E6'),
    right=Side(style='thin', color='E7E6E6'),
    top=Side(style='thin', color='E7E6E6'),
    bottom=Side(style='thin', color='E7E6E6'),
)
_CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

UUID_PATTERN = re.compile(
    r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
)
_ASSET_CODE_PATTERN = re.compile(r'^[A-Z]+\d+$')

# ─── Template generation ─────────────────────────────────────────────

def _generate_template(
    category: str,
    mode: str,
    title: str,
    fields: List[Tuple[str, str, bool]],
    example_data: List[str],
) -> BytesIO:
    """Generic XLSX template generator."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    for col, (field_name, label, _required) in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER

    for col, (field_name, _label, _required) in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col, value="")
        cell.font = _HINT_FONT
        cell.border = _THIN_BORDER
        if field_name == "organization":
            cell.value = "（填写完整路径，节点不存在时会自动创建）"
        elif field_name == "runs_on":
            cell.value = "（主机名称，每行一个）"
        elif field_name == "storage_locations":
            cell.value = "（路径|类型|描述，每行一个）"

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = _CELL_ALIGN
        cell.border = _THIN_BORDER

    for idx in range(len(fields)):
        col_letter = ws.cell(row=1, column=idx + 1).column_letter
        field_name = fields[idx][0]
        if field_name in ('id', 'external_address', 'internal_address'):
            ws.column_dimensions[col_letter].width = 25
        elif field_name == 'credentials':
            ws.column_dimensions[col_letter].width = 20
        elif field_name == 'organization':
            ws.column_dimensions[col_letter].width = 18
        elif field_name == 'storage_locations':
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_category_template(category: str, mode: str) -> Tuple[BytesIO, str]:
    """Generate XLSX template for a given category and mode.

    Replaces the 12 individual generate_*_template() functions.
    Returns (buffer, filename).
    """
    fields = CATEGORY_FIELDS[category][mode]
    example = CATEGORY_EXAMPLES[category][mode]
    title = CATEGORY_EXAMPLES[category][f"title_{mode}"]

    filename_map = {
        ("host", "create"): "主机创建模板.xlsx",
        ("host", "update"): "主机更新模板.xlsx",
        ("network", "create"): "网络设备创建模板.xlsx",
        ("network", "update"): "网络设备更新模板.xlsx",
        ("database", "create"): "数据库创建模板.xlsx",
        ("database", "update"): "数据库更新模板.xlsx",
        ("cloud", "create"): "云服务创建模板.xlsx",
        ("cloud", "update"): "云服务更新模板.xlsx",
        ("web", "create"): "网站服务创建模板.xlsx",
        ("web", "update"): "网站服务更新模板.xlsx",
        ("gpt", "create"): "AI 服务创建模板.xlsx",
        ("gpt", "update"): "AI 服务更新模板.xlsx",
    }
    filename = filename_map[(category, mode)]

    return _generate_template(category, mode, title, fields, example), filename

# Backward-compatible aliases — kept so existing imports still work
generate_host_create_template = lambda: generate_category_template("host", "create")[0]
generate_host_update_template = lambda: generate_category_template("host", "update")[0]
generate_network_create_template = lambda: generate_category_template("network", "create")[0]
generate_network_update_template = lambda: generate_category_template("network", "update")[0]
generate_database_create_template = lambda: generate_category_template("database", "create")[0]
generate_database_update_template = lambda: generate_category_template("database", "update")[0]
generate_cloud_create_template = lambda: generate_category_template("cloud", "create")[0]
generate_cloud_update_template = lambda: generate_category_template("cloud", "update")[0]
generate_web_create_template = lambda: generate_category_template("web", "create")[0]
generate_web_update_template = lambda: generate_category_template("web", "update")[0]
generate_gpt_create_template = lambda: generate_category_template("gpt", "create")[0]
generate_gpt_update_template = lambda: generate_category_template("gpt", "update")[0]

# ─── Helpers ─────────────────────────────────────────────────────────

async def resolve_or_create_organization(db: AsyncSession, path_str: str) -> int:
    """根据路径字符串解析或创建组织节点，返回 org_id。

    路径格式: Default/研发部/服务器组 或 研发部/服务器组（自动补 Default）
    不存在的节点会按层级自动创建。

    Optimized: single batch lookup, then 2 flushes total (IDs + paths)
    regardless of how many new orgs are created.
    """
    parts = [p.strip() for p in path_str.split("/") if p.strip()]
    if not parts:
        raise ValueError("节点路径为空")
    if parts[0] != "Default":
        parts.insert(0, "Default")

    # Batch-lookup: group by name (same name can exist under different parents)
    loaded: Dict[str, List[Organization]] = {}
    result = await db.execute(
        select(Organization).where(Organization.name.in_(parts))
    )
    for org in result.scalars().all():
        loaded.setdefault(org.name, []).append(org)

    # Walk the path, matching by name + parent_id
    current_obj: Optional[Organization] = None
    new_orgs: List[Organization] = []

    for name in parts:
        candidate = None
        for org in loaded.get(name, []):
            if current_obj is None and org.parent_id is None:
                candidate = org
                break
            elif current_obj is not None and org.parent_id == current_obj.id:
                candidate = org
                break
        if candidate is None:
            candidate = Organization(
                name=name,
                parent_id=current_obj.id if current_obj else None,
                level=current_obj.level + 1 if current_obj else 0,
            )
            db.add(candidate)
            new_orgs.append(candidate)
            CreatedOrgs.add(path_str)
        current_obj = candidate

    if new_orgs:
        # First flush: assign IDs
        await db.flush()
        # Set paths
        for org in new_orgs:
            org.path = (
                f"{org.parent_id}/{org.id}" if org.parent_id
                else str(org.id)
            )
        # Second flush: persist paths
        await db.flush()

    if current_obj is None:
        raise ValueError(f"无法解析组织路径: {path_str}")

    return current_obj.id


def parse_status(value) -> Optional[str]:
    """Parse status value from import data.

    Supports:
    - Status enum values: inventory, deploying, running, etc.
    - Legacy Chinese: 启用 → running, 禁用 → deactivated
    - Legacy bool: True → running, False → deactivated
    - Legacy int: 1 → running, 0 → deactivated
    """
    from app.models import AssetStatus

    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if isinstance(value, str):
        v = value.lower().strip()
        for s in AssetStatus:
            if s.value == v:
                return s.value
        if v in ("启用", "是"):
            return AssetStatus.RUNNING.value
        if v in ("禁用", "否"):
            return AssetStatus.DEACTIVATED.value
    elif isinstance(value, bool):
        return AssetStatus.RUNNING.value if value else AssetStatus.DEACTIVATED.value
    elif isinstance(value, (int, float)):
        return AssetStatus.RUNNING.value if value else AssetStatus.DEACTIVATED.value

    return None


def _create_credentials(asset, record: Dict[str, Any], cred_type: str, db):
    """Create Credential objects for an asset from parsed record."""
    creds = record.get("credentials")
    if not creds:
        return
    for cred in creds:
        db.add(Credential(
            asset_id=asset.id,
            username=cred["username"],
            password_encrypted=encrypt_value(cred["password"]),
            credential_type=cred_type,
        ))


async def _replace_credentials(asset_id, record: Dict[str, Any], cred_type: str, db):
    """Delete existing credentials and create new ones from record."""
    await db.execute(delete(Credential).where(Credential.asset_id == asset_id))
    if record.get("credentials"):
        for cred in record["credentials"]:
            db.add(Credential(
                asset_id=asset_id,
                username=cred["username"],
                password_encrypted=encrypt_value(cred["password"]),
                credential_type=cred_type,
            ))


async def _handle_database_relations(asset, record: Dict[str, Any], db):
    """Handle runs_on hosts and storage_locations for database assets."""
    if record.get("runs_on"):
        for host_name in record["runs_on"]:
            host_result = await db.execute(
                select(Asset).where(
                    Asset.category == "host",
                    Asset.name == host_name,
                )
            )
            host = host_result.scalar_one_or_none()
            if host:
                db.add(AssetHostRelation(asset_id=asset.id, host_id=host.id))

    if record.get("storage_locations"):
        for loc in record["storage_locations"]:
            db.add(StorageLocation(
                asset_id=asset.id,
                path=loc["path"],
                path_type=loc["path_type"],
                description=loc.get("description"),
            ))


async def _replace_database_relations(asset_id, record: Dict[str, Any], db):
    """Replace runs_on hosts and storage_locations for database assets."""
    if "runs_on" in record:
        await db.execute(delete(AssetHostRelation).where(AssetHostRelation.asset_id == asset_id))
        if record.get("runs_on"):
            for host_name in record["runs_on"]:
                host_result = await db.execute(
                    select(Asset).where(
                        Asset.category == "host",
                        Asset.name == host_name,
                    )
                )
                host = host_result.scalar_one_or_none()
                if host:
                    db.add(AssetHostRelation(asset_id=asset_id, host_id=host.id))

    if "storage_locations" in record:
        await db.execute(delete(StorageLocation).where(StorageLocation.asset_id == asset_id))
        if record.get("storage_locations"):
            for loc in record["storage_locations"]:
                db.add(StorageLocation(
                    asset_id=asset_id,
                    path=loc["path"],
                    path_type=loc["path_type"],
                    description=loc.get("description"),
                ))

# ─── Generic batch operations ────────────────────────────────────────

async def batch_create_assets(
    category: str,
    records: List[Dict[str, Any]],
    db: AsyncSession,
    user_id: Optional[int] = None,
) -> Tuple[int, List[Dict[str, Any]]]:
    """Batch create assets for any category.

    Replaces the 6 individual batch_create_* functions.
    Uses a single query to check all duplicate names at once.
    """
    cred_type = CATEGORY_CRED_TYPE[category]
    create_fields = CATEGORY_CREATE_FIELDS[category]
    success_count = 0
    failed_records = []

    # Single batch query for duplicate names
    names = [r["name"] for r in records]
    existing_result = await db.execute(
        select(Asset.name).where(
            Asset.category == category,
            Asset.name.in_(names),
        )
    )
    existing_names = {row.name for row in existing_result.fetchall()}

    for record in records:
        try:
            if record["name"] in existing_names:
                failed_records.append({
                    "name": record["name"],
                    "error": f"资产名称'{record['name']}'已存在",
                })
                continue

            kwargs: Dict[str, Any] = {
                "name": record["name"],
                "category": category,
                "created_by_id": user_id,
            }
            for field in create_fields:
                val = record.get(field)
                if val is not None:
                    kwargs[field] = val

            # OOB password encryption (host only)
            if record.get("oob_password"):
                kwargs["oob_password_encrypted"] = encrypt_value(record["oob_password"])

            # Database: version → extra_data
            if category == "database" and record.get("version"):
                kwargs["extra_data"] = {"version": record["version"]}

            asset = Asset(**kwargs)
            db.add(asset)
            await db.flush()

            status = parse_status(record.get("status"))
            if status:
                asset.status = status

            _create_credentials(asset, record, cred_type, db)

            if category == "database":
                await _handle_database_relations(asset, record, db)

            success_count += 1
        except Exception as e:
            failed_records.append({"name": record.get("name"), "error": str(e)})

    if success_count > 0:
        await db.commit()

    return success_count, failed_records


async def batch_update_assets(
    category: str,
    records: List[Dict[str, Any]],
    db: AsyncSession,
) -> Tuple[int, List[Dict[str, Any]]]:
    """Batch update assets for any category.

    Replaces the 6 individual batch_update_* functions.
    """
    cred_type = CATEGORY_CRED_TYPE[category]
    update_fields = CATEGORY_UPDATE_FIELDS[category]
    success_count = 0
    failed_records = []

    for record in records:
        try:
            result = await db.execute(select(Asset).where(Asset.id == record["id"]))
            asset = result.scalar_one_or_none()
            if not asset:
                failed_records.append({"id": record.get("id"), "error": "资产不存在"})
                continue

            for field in update_fields:
                if record.get(field) is not None:
                    setattr(asset, field, record[field])

            if record.get("organization_id"):
                asset.organization_id = record["organization_id"]

            # Host: OOB password
            if category == "host" and record.get("oob_password") is not None:
                if record["oob_password"]:
                    asset.oob_password_encrypted = encrypt_value(record["oob_password"])
                else:
                    asset.oob_password_encrypted = None

            # Database: version → extra_data
            if category == "database" and record.get("version"):
                current_extra = asset.extra_data or {}
                asset.extra_data = {**current_extra, "version": record["version"]}

            status = parse_status(record.get("status"))
            if status:
                asset.status = status

            if "credentials" in record:
                await _replace_credentials(asset.id, record, cred_type, db)

            if category == "database":
                await _replace_database_relations(asset.id, record, db)

            success_count += 1
        except Exception as e:
            failed_records.append({"id": record.get("id"), "error": str(e)})

    if success_count > 0:
        await db.commit()

    return success_count, failed_records

# ─── Backward-compatible aliases ─────────────────────────────────────

async def batch_create_hosts(records, db, user_id=None):
    return await batch_create_assets("host", records, db, user_id)

async def batch_update_hosts(records, db):
    return await batch_update_assets("host", records, db)

async def batch_create_networks(records, db, user_id=None):
    return await batch_create_assets("network", records, db, user_id)

async def batch_update_networks(records, db):
    return await batch_update_assets("network", records, db)

async def batch_create_databases(records, db, user_id=None):
    return await batch_create_assets("database", records, db, user_id)

async def batch_update_databases(records, db):
    return await batch_update_assets("database", records, db)

async def batch_create_clouds(records, db, user_id=None):
    return await batch_create_assets("cloud", records, db, user_id)

async def batch_update_clouds(records, db):
    return await batch_update_assets("cloud", records, db)

async def batch_create_webs(records, db, user_id=None):
    return await batch_create_assets("web", records, db, user_id)

async def batch_update_webs(records, db):
    return await batch_update_assets("web", records, db)

async def batch_create_gpts(records, db, user_id=None):
    return await batch_create_assets("gpt", records, db, user_id)

async def batch_update_gpts(records, db):
    return await batch_update_assets("gpt", records, db)

# ─── Parse import file ───────────────────────────────────────────────

async def _load_relevant_orgs(db: AsyncSession, org_values: List[str]) -> Tuple[
    Dict[str, int],  # org_name_map: name → id
    Dict[str, int],  # org_path_map: path → id
    Dict[str, int],  # org_display_path_map: display_path → id
]:
    """Load only organizations referenced by import data and their ancestors.

    Avoids loading the full organizations table when most orgs are irrelevant.
    """
    if not org_values:
        return {}, {}, {}

    # Phase 1: find orgs matching values, collect ancestor IDs
    names_set = set(org_values)
    all_orgs: Dict[int, Organization] = {}
    id_to_name: Dict[int, str] = {}

    while names_set:
        result = await db.execute(
            select(Organization).where(Organization.name.in_(names_set))
        )
        new_names: set[str] = set()
        for org in result.scalars().all():
            if org.id not in all_orgs:
                all_orgs[org.id] = org
                id_to_name[org.id] = org.name
                if org.parent_id and org.parent_id not in all_orgs:
                    # Look up parent name from already-loaded orgs
                    parent_name = id_to_name.get(org.parent_id)
                    if parent_name:
                        new_names.add(parent_name)
        names_set = new_names - set(o.name for o in all_orgs.values())

    # Phase 2: load any missing ancestors by ID
    missing_ids = set()
    for org in all_orgs.values():
        if org.parent_id and org.parent_id not in all_orgs:
            missing_ids.add(org.parent_id)
    while missing_ids:
        result = await db.execute(
            select(Organization).where(Organization.id.in_(missing_ids))
        )
        new_missing: set[int] = set()
        for org in result.scalars().all():
            if org.id not in all_orgs:
                all_orgs[org.id] = org
                id_to_name[org.id] = org.name
                if org.parent_id and org.parent_id not in all_orgs:
                    new_missing.add(org.parent_id)
        missing_ids = new_missing

    # Phase 3: build lookup maps
    org_name_map = {org.name: org.id for org in all_orgs.values()}
    org_path_map = {org.path: org.id for org in all_orgs.values() if org.path}

    org_display_path_map: Dict[str, int] = {}

    for org in all_orgs.values():
        if org.path:
            name_parts = [id_to_name[int(p)] for p in org.path.split('/') if int(p) in id_to_name]
            if name_parts:
                dws = ' / '.join(name_parts)
                dwo = '/'.join(name_parts)
                org_display_path_map[dws] = org.id
                org_display_path_map[dwo] = org.id
                org_display_path_map[f'Default / {dws}'] = org.id
                org_display_path_map[f'Default/{dwo}'] = org.id
        org_display_path_map[org.name] = org.id

    return org_name_map, org_path_map, org_display_path_map


async def parse_import_file(
    file_content: bytes,
    category: str,
    mode: str,
    db: AsyncSession,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Parse XLSX file and validate data.

    Returns: (valid_records, error_records)
    Side effect: CreatedOrgs is cleared at start, populated with
        auto-created org paths during parsing for audit reporting.
    """
    CreatedOrgs.clear()
    wb = load_workbook(BytesIO(file_content))
    ws = wb.active

    if category not in CATEGORY_FIELDS:
        raise ValueError(f"不支持的资产类型：{category}")

    fields = CATEGORY_FIELDS[category][mode]

    valid_records = []
    error_records = []

    # Header mapping
    header_row = ws.iter_rows(min_row=1, max_row=1, values_only=True).__next__()
    header_map = {}
    expected_labels = {field_label.lstrip('*').strip() for _, field_label, _ in fields}

    for col_idx, header in enumerate(header_row):
        if header:
            label_key = str(header).strip().replace('*', '').replace('*', '').strip()
            header_map[label_key] = col_idx

    # Validate headers
    missing_headers = [h for h in expected_labels if h not in header_map]
    unrecognized_headers = [h for h in header_map.keys() if h not in expected_labels]

    if missing_headers:
        missing_display = []
        for h in missing_headers:
            for field_name, label, required in fields:
                if label.lstrip('*') == h and required:
                    missing_display.append(f"*{h}")
                    break
            else:
                missing_display.append(h)

        error_msg = f"表头缺失字段：{', '.join(missing_display)}"
        if unrecognized_headers:
            extra = ', '.join(unrecognized_headers[:3])
            if len(unrecognized_headers) > 3:
                extra += '...'
            error_msg += f"（发现未识别列：{extra}）"

        return [], [{
            "row": 1,
            "errors": [error_msg],
            "data": {"提示": f"请检查是否使用了正确的{CATEGORY_NAMES.get(category, category)}{'创建' if mode == 'create' else '更新'}模板"},
        }]

    # Scan data rows for organization values (before loading from DB)
    org_col_idx = None
    for field_name, label, _ in fields:
        if field_name == "organization":
            org_col_idx = header_map.get(label.lstrip('*').strip())
            break

    raw_org_values: List[str] = []
    if org_col_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if org_col_idx < len(row) and row[org_col_idx]:
                val = str(row[org_col_idx]).strip()
                if val:
                    raw_org_values.append(val)

    # Extract individual org names from path-like values
    # (e.g., "Default / 研发部 / 服务器组" → ["研发部", "服务器组"])
    org_values: List[str] = []
    for val in raw_org_values:
        org_values.append(val)
        # Try splitting on common path separators
        for sep in ('/', ' / '):
            if sep in val:
                org_values.extend(n.strip() for n in val.split(sep) if n.strip() != 'Default')
                break

    # Load only relevant organizations on demand
    org_name_map, org_path_map, org_display_path_map = await _load_relevant_orgs(
        db, org_values
    )

    # Data rows
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue

        record = {}
        errors = []

        for field_name, label, required in fields:
            label_key = label.lstrip('*').strip()
            col_idx = header_map.get(label_key)

            if col_idx is not None and col_idx < len(row):
                value = row[col_idx]
            else:
                value = None

            if value is not None:
                value = str(value).strip() if value else None
                if value == '':
                    value = None

            if required and not value:
                errors.append(f"{label}为必填项")

            if field_name == "organization" and value:
                org_id = (
                    org_display_path_map.get(value)
                    or org_path_map.get(value)
                    or org_name_map.get(value)
                )
                if not org_id:
                    try:
                        org_id = await resolve_or_create_organization(db, value)
                        record["organization_id"] = org_id
                    except ValueError:
                        if _ASSET_CODE_PATTERN.match(value):
                            errors.append(f"组织节点'{value}'不存在（提示：该值看起来像资产编号，请检查是否误填至节点列）")
                        else:
                            errors.append(f"组织节点'{value}'不存在且无法自动创建")
                else:
                    record["organization_id"] = org_id

            elif field_name == "credentials" and value:
                credentials_list = []
                for line_num, line in enumerate(value.split('\n'), 1):
                    line = line.strip()
                    if not line:
                        continue
                    if ':' not in line:
                        errors.append(f"用户名密码格式错误（第{line_num}行）：'{line}'，应为 username:password 格式")
                        continue
                    ci = line.index(':')
                    username = line[:ci].strip()
                    password = line[ci + 1:].strip()
                    if not username:
                        errors.append(f"用户名密码格式错误（第{line_num}行）：用户名为空")
                        continue
                    if not password:
                        errors.append(f"用户名密码格式错误（第{line_num}行）：密码为空")
                        continue
                    credentials_list.append({"username": username, "password": password})
                if credentials_list:
                    record["credentials"] = credentials_list

            elif field_name in ("oob", "oob_username", "oob_password"):
                if value:
                    if field_name == "oob":
                        record["oob_address"] = value
                    else:
                        record[field_name] = value

            elif field_name == "runs_on" and value:
                host_names = [h.strip() for h in value.split('\n') if h.strip()]
                if host_names:
                    record["runs_on"] = host_names

            elif field_name == "storage_locations" and value:
                locations = []
                for line_num, line in enumerate(value.split('\n'), 1):
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split('|')
                    if len(parts) < 2:
                        errors.append(f"存储位置格式错误（第{line_num}行）：'{line}'，应为 路径|类型|描述 格式")
                        continue
                    path_val = parts[0].strip()
                    path_type = parts[1].strip()
                    description = parts[2].strip() if len(parts) > 2 else None
                    if not path_val:
                        errors.append(f"存储位置格式错误（第{line_num}行）：路径为空")
                        continue
                    if path_type not in ("data", "log", "backup", "temp"):
                        errors.append(f"存储位置类型错误（第{line_num}行）：'{path_type}'，应为 data/log/backup/temp")
                        continue
                    locations.append({
                        "path": path_val,
                        "path_type": path_type,
                        "description": description,
                    })
                if locations:
                    record["storage_locations"] = locations

            elif field_name == "id" and value:
                value = str(value).strip()
                if not value:
                    errors.append("ID 不能为空")
                elif mode == "update" and not UUID_PATTERN.match(value.lower()):
                    errors.append(f"ID 格式无效：'{value}'，应为 UUID 格式")
                record["id"] = value

            elif field_name != "organization":
                record[field_name] = value

        if errors:
            context = {k: v for k, v in record.items() if v is not None and k in ('name', 'id', 'asset_code')}
            first_non_empty = next((str(v) for v in row if v), None)
            if first_non_empty:
                context['row_data_preview'] = first_non_empty[:50]
            error_records.append({
                "row": row_num,
                "errors": errors,
                "data": context or {"row_content": " | ".join(str(v) for v in row[:5] if v)},
            })
        else:
            valid_records.append(record)

    return valid_records, error_records
