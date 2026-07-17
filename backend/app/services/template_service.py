"""
Asset Import Template Service
XLSX template generation for asset bulk import/update
"""
from io import BytesIO
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ─── Field definitions ───────────────────────────────────────────────
# Format: (field_name, display_label, is_required)

HOST_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "外网地址", False),
    ("internal_address", "*内网地址", True),
    ("credentials", "*用户名密码", True),
    ("cpu", "CPU", False),
    ("memory", "内存", False),
    ("system_disk", "系统盘", False),
    ("data_disk", "数据盘", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("oob", "OOB 地址", False),
    ("oob_username", "OOB 用户名", False),
    ("oob_password", "OOB 密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

HOST_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("cpu", "CPU", False),
    ("memory", "内存", False),
    ("system_disk", "系统盘", False),
    ("data_disk", "数据盘", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("oob", "OOB 地址", False),
    ("oob_username", "OOB 用户名", False),
    ("oob_password", "OOB 密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

NETWORK_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("device_type", "*设备类型", True),
    ("vendor", "*厂商", True),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("status", "*状态", True),
    ("notes", "描述", False),
]

NETWORK_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("device_type", "设备类型", False),
    ("vendor", "厂商", False),
    ("model", "型号", False),
    ("serial_number", "序列号", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

DATABASE_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("db_type", "*数据库类型", True),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("runs_on", "运行于", False),
    ("storage_locations", "存储位置", False),
    ("version", "版本", False),
    ("namespace", "命名空间", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

DATABASE_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("db_type", "数据库类型", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("runs_on", "运行于", False),
    ("storage_locations", "存储位置", False),
    ("version", "版本", False),
    ("namespace", "命名空间", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CLOUD_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CLOUD_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

WEB_CREATE_FIELDS = [
    ("name", "*资产名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "*外网地址", True),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

WEB_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "资产名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

GPT_CREATE_FIELDS = [
    ("name", "*名称", True),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "*平台", True),
    ("external_address", "*外网地址", True),
    ("internal_address", "内网地址", False),
    ("credentials", "*用户名密码", True),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

GPT_UPDATE_FIELDS = [
    ("id", "*ID", True),
    ("name", "名称", False),
    ("asset_code", "资产编号", False),
    ("organization", "节点", False),
    ("platform", "平台", False),
    ("external_address", "外网地址", False),
    ("internal_address", "内网地址", False),
    ("credentials", "用户名密码", False),
    ("applicant", "申请人", False),
    ("status", "状态", False),
    ("notes", "描述", False),
]

CATEGORY_FIELDS = {
    "host": {"create": HOST_CREATE_FIELDS, "update": HOST_UPDATE_FIELDS},
    "network": {"create": NETWORK_CREATE_FIELDS, "update": NETWORK_UPDATE_FIELDS},
    "database": {"create": DATABASE_CREATE_FIELDS, "update": DATABASE_UPDATE_FIELDS},
    "cloud": {"create": CLOUD_CREATE_FIELDS, "update": CLOUD_UPDATE_FIELDS},
    "web": {"create": WEB_CREATE_FIELDS, "update": WEB_UPDATE_FIELDS},
    "gpt": {"create": GPT_CREATE_FIELDS, "update": GPT_UPDATE_FIELDS},
}

# ─── Template example data ───────────────────────────────────────────

CATEGORY_EXAMPLES = {
    "host": {
        "title_create": "主机导入模板",
        "title_update": "主机更新模板",
        "create": [
            "test-server-01", "CI001", "Default/研发部/服务器组", "Linux",
            "192.168.1.100", "10.0.0.100", "admin:123456", "8 核", "16GB",
            "500GB SSD", "2TB HDD", "Dell R740", "SN123456", "192.168.1.200",
            "admin", "admin123", "张三", "启用", "测试服务器 - 开发环境",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "test-server-01", "CI001",
            "Default/研发部/服务器组", "Linux", "192.168.1.100", "10.0.0.100",
            "admin:123456", "8 核", "16GB", "500GB SSD", "2TB HDD", "Dell R740",
            "SN123456", "192.168.1.200", "admin", "admin123", "张三", "启用",
            "测试服务器 - 开发环境",
        ],
    },
    "network": {
        "title_create": "网络设备导入模板",
        "title_update": "网络设备更新模板",
        "create": [
            "Core-SW-01", "NW001", "Default/研发部/网络设备", "交换机", "Cisco",
            "C9300-48P", "FCW1234D001", "10.0.0.1", "192.168.1.1",
            "admin:cisco123\nnetadmin:netpass123", "启用", "核心交换机 - 生产环境",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "Core-SW-01", "NW001",
            "Default/研发部/网络设备", "交换机", "Cisco", "C9300-48P", "FCW1234D001",
            "10.0.0.1", "192.168.1.1", "admin:cisco123\nnetadmin:netpass123",
            "启用", "核心交换机 - 生产环境",
        ],
    },
    "database": {
        "title_create": "数据库导入模板",
        "title_update": "数据库更新模板",
        "create": [
            "MySQL-Prod-01", "DB001", "Default/研发部/数据库", "RDS", "MySQL",
            "", "192.168.1.100:3306", "root:mysqlroot123\napp:apppass",
            "web-server-01", "/var/lib/mysql|data|主数据目录", "8.0.32", "main",
            "张三", "True", "生产环境主数据库",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "MySQL-Prod-01", "DB001",
            "Default/研发部/数据库", "RDS", "MySQL", "", "192.168.1.100:3306",
            "root:mysqlroot123\napp:apppass", "web-server-01",
            "/var/lib/mysql|data|主数据目录", "8.0.32", "main", "张三",
            "running", "生产环境主数据库",
        ],
    },
    "cloud": {
        "title_create": "云服务导入模板",
        "title_update": "云服务更新模板",
        "create": [
            "AWS-Prod-Account", "CL001", "Default/研发部/云服务", "AWS",
            "https://aws.amazon.com", "10.0.0.1",
            "AKIAIOSFODNN7EXAMPLE:wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
            "True", "生产环境 AWS 账号",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "AWS-Prod-Account", "CL001",
            "Default/研发部/云服务", "AWS", "https://aws.amazon.com", "10.0.0.1",
            "AKIAIOSFODNN7EXAMPLE:wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
            "启用", "生产环境 AWS 账号",
        ],
    },
    "web": {
        "title_create": "网站服务导入模板",
        "title_update": "网站服务更新模板",
        "create": [
            "Jira", "WB001", "Default/研发部/应用系统", "Nginx",
            "https://jira.example.com", "http://192.168.1.100:8080",
            "admin:jiraadmin\nreadonly:readonly123", "张三", "True",
            "项目管理平台",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "Jira", "WB001",
            "Default/研发部/应用系统", "Nginx", "https://jira.example.com",
            "http://192.168.1.100:8080",
            "admin:jiraadmin\nreadonly:readonly123", "项目管理平台",
        ],
    },
    "gpt": {
        "title_create": "AI 服务导入模板",
        "title_update": "AI 服务更新模板",
        "create": [
            "OpenAI-API", "AI001", "Default/研发部/AI 服务", "OpenAI",
            "https://api.openai.com/v1", "",
            "sk-key:sk-abc123xyz\nclue-key:sk-clue456", "张三", "True",
            "AI 服务 API",
        ],
        "update": [
            "56c4d4cd-42ba-4397-abfa-36ecba64af13", "OpenAI-API", "AI001",
            "Default/研发部/AI 服务", "OpenAI", "https://api.openai.com/v1", "",
            "sk-key:sk-abc123xyz\nclue-key:sk-clue456", "张三", "running",
            "AI 服务 API",
        ],
    },
}

# ─── Shared openpyxl styles ──────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
_HINT_FONT = Font(italic=True, color="666666", size=9)
_THIN_BORDER = Border(
    left=Side(style='thin', color='E7E6E6'),
    right=Side(style='thin', color='E7E6E6'),
    top=Side(style='thin', color='E7E6E6'),
    bottom=Side(style='thin', color='E7E6E6'),
)
_CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ─── Template generation ─────────────────────────────────────────────

def _generate_template(
    category: str,
    mode: str,
    title: str,
    fields: List[Tuple[str, str, bool]],
    example_data: List[str],
) -> BytesIO:
    """Generic XLSX template generator."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    for col, (field_name, label, _required) in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER

    for col, (field_name, _label, _required) in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col, value="")
        cell.font = _HINT_FONT
        cell.border = _THIN_BORDER
        if field_name == "organization":
            cell.value = "（填写完整路径，节点不存在时会自动创建）"
        elif field_name == "runs_on":
            cell.value = "（主机名称，每行一个）"
        elif field_name == "storage_locations":
            cell.value = "（路径|类型|描述，每行一个）"

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = _CELL_ALIGN
        cell.border = _THIN_BORDER

    for idx in range(len(fields)):
        col_letter = ws.cell(row=1, column=idx + 1).column_letter
        field_name = fields[idx][0]
        if field_name in ('id', 'external_address', 'internal_address'):
            ws.column_dimensions[col_letter].width = 25
        elif field_name == 'credentials':
            ws.column_dimensions[col_letter].width = 20
        elif field_name == 'organization':
            ws.column_dimensions[col_letter].width = 18
        elif field_name == 'storage_locations':
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_category_template(category: str, mode: str) -> Tuple[BytesIO, str]:
    """Generate XLSX template for a given category and mode.

    Replaces the 12 individual generate_*_template() functions.
    Returns (buffer, filename).
    """
    fields = CATEGORY_FIELDS[category][mode]
    example = CATEGORY_EXAMPLES[category][mode]
    title = CATEGORY_EXAMPLES[category][f"title_{mode}"]

    filename_map = {
        ("host", "create"): "主机创建模板.xlsx",
        ("host", "update"): "主机更新模板.xlsx",
        ("network", "create"): "网络设备创建模板.xlsx",
        ("network", "update"): "网络设备更新模板.xlsx",
        ("database", "create"): "数据库创建模板.xlsx",
        ("database", "update"): "数据库更新模板.xlsx",
        ("cloud", "create"): "云服务创建模板.xlsx",
        ("cloud", "update"): "云服务更新模板.xlsx",
        ("web", "create"): "网站服务创建模板.xlsx",
        ("web", "update"): "网站服务更新模板.xlsx",
        ("gpt", "create"): "AI 服务创建模板.xlsx",
        ("gpt", "update"): "AI 服务更新模板.xlsx",
    }
    filename = filename_map[(category, mode)]

    return _generate_template(category, mode, title, fields, example), filename
