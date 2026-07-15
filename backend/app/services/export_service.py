"""
Asset Export Service
Handles Excel and CSV export functionality
"""
from io import BytesIO, StringIO
import csv
from typing import List, Dict, Any, Tuple
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from app.models import AssetStatus
from app.core.asset_categories import ASSET_CATEGORY_LABELS

# Column definitions for export by asset category
CATEGORY_COLUMNS: Dict[str, List[tuple]] = {
    "host": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("platform", "平台"),
        ("organization_name", "节点"),
        ("model", "型号"),
        ("serial_number", "序列号"),
        ("cpu", "CPU"),
        ("memory", "内存"),
        ("system_disk", "系统盘"),
        ("data_disk", "数据盘"),
        ("credentials", "用户名密码"),
        ("oob", "OOB 地址"),
        ("oob_username", "OOB 用户名"),
        ("oob_password", "OOB 密码"),
        ("applicant", "申请人"),
        ("owner_name", "负责人"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
    "network": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("device_type", "设备类型"),
        ("vendor", "厂商"),
        ("model", "型号"),
        ("serial_number", "序列号"),
        ("organization_name", "节点"),
        ("credentials", "用户名密码"),
        ("owner_name", "负责人"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
    "database": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("platform", "平台"),
        ("db_type", "数据库类型"),
        ("version", "版本"),
        ("namespace", "命名空间"),
        ("runs_on", "运行于"),
        ("storage_locations", "存储位置"),
        ("organization_name", "节点"),
        ("applicant", "申请人"),
        ("owner_name", "负责人"),
        ("credentials", "用户名密码"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
    "cloud": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("platform", "平台"),
        ("organization_name", "节点"),
        ("owner_name", "负责人"),
        ("credentials", "用户名密码"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
    "web": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("platform", "平台"),
        ("organization_name", "节点"),
        ("applicant", "申请人"),
        ("owner_name", "负责人"),
        ("credentials", "用户名密码"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
    "gpt": [
        ("id", "ID"),
        ("name", "资产名称"),
        ("asset_code", "资产编号"),
        ("internal_address", "内网地址"),
        ("external_address", "外网地址"),
        ("platform", "平台"),
        ("organization_name", "节点"),
        ("owner_name", "负责人"),
        ("credentials", "用户名密码"),
        ("notes", "描述"),
        ("creator_name", "创建者"),
        ("status", "状态"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ],
}

# Default columns for mixed exports (all categories)
DEFAULT_COLUMNS = [
    ("id", "ID"),
    ("name", "资产名称"),
    ("asset_code", "资产编号"),
    ("category", "资产类型"),
    ("internal_address", "内网地址"),
    ("external_address", "外网地址"),
    ("platform", "平台"),
    ("organization_name", "节点"),
    ("device_type", "设备类型"),
    ("vendor", "厂商"),
    ("db_type", "数据库类型"),
    ("model", "型号"),
    ("serial_number", "序列号"),
    ("cpu", "CPU"),
    ("memory", "内存"),
    ("system_disk", "系统盘"),
    ("data_disk", "数据盘"),
    ("credentials", "用户名密码"),
    ("oob", "OOB 地址"),
    ("oob_username", "OOB 用户名"),
    ("oob_password", "OOB 密码"),
    ("applicant", "申请人"),
    ("owner_name", "负责人"),
    ("notes", "描述"),
    ("creator_name", "创建者"),
    ("status", "状态"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
]

UTC8 = timezone(timedelta(hours=8))
UTC8_FMT = "%Y-%m-%d %H:%M:%S"


# Module-level status label mapping — reused across all exports
STATUS_LABELS: Dict[str, str] = {
    AssetStatus.INVENTORY: "库存",
    AssetStatus.DEPLOYING: "部署中",
    AssetStatus.RUNNING: "运行中",
    AssetStatus.MAINTENANCE: "维护中",
    AssetStatus.DEACTIVATED: "停用",
    AssetStatus.PENDING_SCRAP: "待报废",
    AssetStatus.SCRAPPED: "已报废",
    AssetStatus.RETURNED: "已退还",
}

# Column width defaults
COLUMN_WIDTHS = {
    "id": 10,
    "name": 20,
    "asset_code": 15,
    "category": 12,
    "internal_address": 20,
    "external_address": 20,
    "organization_name": 20,
    "device_type": 12,
    "vendor": 15,
    "model": 20,
    "serial_number": 20,
    "cpu": 12,
    "memory": 12,
    "system_disk": 15,
    "data_disk": 15,
    "credentials": 25,
    "oob": 20,
    "oob_username": 15,
    "oob_password": 15,
    "applicant": 12,
    "runs_on": 20,
    "storage_locations": 25,
    "owner_name": 15,
    "notes": 30,
    "creator_name": 12,
    "status": 12,
    "created_at": 20,
    "updated_at": 20,
}


def format_field_value(field: str, asset: Dict[str, Any]) -> Any:
    """Format a single field value for export. Shared by Excel and CSV writers."""
    value = asset.get(field)

    if field == "category" and value:
        return ASSET_CATEGORY_LABELS.get(value, value)
    if field == "status":
        return STATUS_LABELS.get(value, value) if value else ""
    if field == "vendor" and value is None:
        return asset.get("platform")
    if field in ("created_at", "updated_at") and value:
        if isinstance(value, datetime):
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
            value = value.astimezone(UTC8)
            return value.strftime(UTC8_FMT)
        if isinstance(value, str):
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt.astimezone(UTC8).strftime(UTC8_FMT)
            except Exception:
                pass
        return str(value)
    if field == "credentials":
        creds = asset.get("credentials", [])
        if creds:
            return "\n".join(f"{c.get('username')}:{c.get('password', '')}" for c in creds if c.get('username'))
        return ""
    if field == "oob":
        return asset.get("oob_address")
    if field == "version":
        return (asset.get("extra_data") or {}).get("version")
    if field == "runs_on":
        hosts = asset.get("runs_on_hosts", [])
        return ", ".join(h.get("name", "") for h in hosts) if hosts else ""
    if field == "storage_locations":
        locs = asset.get("storage_locations", [])
        return "\n".join(f"{l.get('path_type', '')}:{l.get('path', '')}" for l in locs) if locs else ""
    if field == "extra_data":
        return ""
    return value


async def _export_excel_stream(
    data_gen,  # async generator yielding Dict[str, Any]
    export_columns: List[tuple],
) -> Tuple[BytesIO, int]:
    """Stream asset data into an Excel file using write_only mode.

    Avoids holding all asset data in memory at once.
    Note: write_only mode does not support cell-level styling.
    Returns: (BytesIO, row_count)
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # Write headers
    header_row = [label for _, label in export_columns]
    ws.append(header_row)

    # Column widths
    for col_idx, (field, _) in enumerate(export_columns, 1):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(field, 15)

    # Write data rows from generator
    count = 0
    async for asset in data_gen:
        row = []
        for field, _ in export_columns:
            if field == "extra_data":
                row.append("")
                continue
            value = format_field_value(field, asset)
            row.append(value if value else "")
        ws.append(row)
        count += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, count


async def _export_csv_stream(
    data_gen,  # async generator yielding Dict[str, Any]
    export_columns: List[tuple],
) -> Tuple[BytesIO, int]:
    """Stream asset data into a CSV file with UTF-8 BOM.

    Returns: (BytesIO, row_count)
    """
    string_buffer = StringIO()
    writer = csv.writer(string_buffer, lineterminator='\n')

    # Header
    writer.writerow([label for _, label in export_columns])

    count = 0
    async for asset in data_gen:
        row = []
        for field, _ in export_columns:
            if field == "extra_data":
                row.append("")
                continue
            value = format_field_value(field, asset)
            row.append(value if value else "")
        writer.writerow(row)
        count += 1

    csv_content = string_buffer.getvalue()
    string_buffer.close()

    content = b'\xef\xbb\xbf' + csv_content.encode('utf-8')
    return BytesIO(content), count


